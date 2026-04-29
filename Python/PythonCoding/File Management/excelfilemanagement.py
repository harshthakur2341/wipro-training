from openpyxl import Workbook,load_workbook
import os
def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name","Age"])
    sheet.append(["harsh","22"])
    sheet.append(["harshista", "42"])
    workbook.save(filename)

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only= True):
        print(f"Name:  {row[0]},Age: {row[1]}")



filename = "data.xlsx"
write_excel(filename)
print("Data read from Excel File: ")
read_excel(filename)
